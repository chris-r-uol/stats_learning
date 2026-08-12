"""
Build the two student-facing Excel files from the generated CSVs.

  data/xlsx/transport-stats-data.xlsx
      One sheet per dataset.  The travel time sheet deliberately places data in
      B4:B58 so the cell references quoted throughout the tutorials
      (=AVERAGE(B4:B58), =STDEV.S(B4:B58) ...) are literally correct.

  workbooks/hypothesis-test-workbench.xlsx
      One sheet per test.  Yellow cells are inputs; everything else is a live
      formula.  Each sheet computes the critical-value route and the p-value
      route side by side and states a plain-English verdict, because the two
      routes agreeing is the thing students most need to internalise.

Run:  python tools/build_workbooks.py
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
CSV_DIR = ROOT / "data" / "csv"
XLSX_DIR = ROOT / "data" / "xlsx"
WB_DIR = ROOT / "workbooks"

HEAD_FILL = PatternFill("solid", fgColor="1F4E5F")
HEAD_FONT = Font(color="FFFFFF", bold=True)
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
RESULT_FILL = PatternFill("solid", fgColor="DEEAF6")
VERDICT_FILL = PatternFill("solid", fgColor="E2EFDA")
TITLE_FONT = Font(bold=True, size=14, color="1F4E5F")
NOTE_FONT = Font(italic=True, size=9, color="595959")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def read_csv(name: str) -> tuple[list[str], list[list]]:
    with open(CSV_DIR / name) as fh:
        rows = list(csv.reader(fh))
    header, body = rows[0], rows[1:]
    out = []
    for r in body:
        out.append([_num(v) for v in r])
    return header, out


def _num(v: str):
    try:
        f = float(v)
        return int(f) if f.is_integer() else f
    except ValueError:
        return v


def style_header(ws, row: int, ncols: int, start: int = 1) -> None:
    for c in range(start, start + ncols):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX


def autosize(ws, width: int = 16) -> None:
    for c in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(c)].width = width


# ---------------------------------------------------------------------------
# Data workbook
# ---------------------------------------------------------------------------

def build_data_workbook() -> None:
    wb = Workbook()

    # -- Read me ------------------------------------------------------------
    ws = wb.active
    ws.title = "Read Me"
    ws["A1"] = "Transport Data Collection and Analysis - Tutorial Data"
    ws["A1"].font = TITLE_FONT
    notes = [
        "",
        "One sheet per dataset. The sheet name tells you which tutorial uses it.",
        "",
        "Sheet                     Used by",
        "T1 Travel Times           Tutorial 1, Questions 2 and 6",
        "T1 Histogram Guide        Tutorial 1, Questions 3, 4 and 5",
        "T3 Fuel Consumption       Tutorial 3, Question 3",
        "T3 HGV Weights            Tutorial 3, Questions 4 and 5",
        "T4 Junction Waiting       Tutorial 4, Question 5",
        "T5 Taxi Queue             Tutorial 5, Question 1",
        "T5 Car Ownership          Tutorial 5, Question 2",
        "T5 Accidents Weather      Tutorial 5, Question 3",
        "T5 Reaction Distance      Tutorial 5, Question 4",
        "",
        "On the travel time sheet the observations sit in B4:B58, so the cell",
        "references used in the tutorial worked solutions apply exactly as written.",
        "",
        "The same data is available as CSV in data/csv/ if you prefer R or Python.",
    ]
    for i, line in enumerate(notes, start=2):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 78

    # -- Travel times: data must land in B4:B58 -----------------------------
    ws = wb.create_sheet("T1 Travel Times")
    ws["A1"] = "Car travel time survey"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Time taken to travel between two locations, in minutes."
    ws["A2"].font = NOTE_FONT
    ws["A3"], ws["B3"] = "Observation", "Travel time (min)"
    style_header(ws, 3, 2)
    _, rows = read_csv("travel-times.csv")
    for i, (obs, val) in enumerate(rows):
        ws.cell(row=4 + i, column=1, value=obs).border = BOX
        ws.cell(row=4 + i, column=2, value=val).border = BOX
    ws["D4"] = "Data occupies B4:B58 - the ranges quoted in the tutorials."
    ws["D4"].font = NOTE_FONT
    autosize(ws, 18)

    # -- Histogram guide ----------------------------------------------------
    ws = wb.create_sheet("T1 Histogram Guide")
    ws["A1"] = "Travel time count data, class width 5"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Columns C, D and E are for you to complete in Questions 3 and 4. "
                "See the tutorial for the definitions.")
    ws["A2"].font = NOTE_FONT
    headers = ["Bin", "Count", "Relative frequency",
               "Relative frequency density", "Cumulative relative frequency"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)
    style_header(ws, 3, len(headers))
    bins = [(0, 5), (5, 10), (10, 15), (15, 20), (20, 25), (25, 30), (30, 35),
            (35, 40), (40, 45), (45, 50), (50, 55), (55, 60), (60, 65)]
    counts = [0, 0, 16, 16, 4, 5, 7, 4, 2, 0, 0, 0, 1]
    for i, ((lo, hi), n) in enumerate(zip(bins, counts)):
        r = 4 + i
        ws.cell(row=r, column=1, value=f"({lo},{hi}]").border = BOX
        ws.cell(row=r, column=2, value=n).border = BOX
        for c in (3, 4, 5):
            ws.cell(row=r, column=c).border = BOX
            ws.cell(row=r, column=c).fill = INPUT_FILL
    r = 4 + len(bins)
    ws.cell(row=r, column=1, value="Total").font = BOLD
    ws.cell(row=r, column=2, value=f"=SUM(B4:B{r-1})").font = BOLD
    autosize(ws, 24)
    ws.column_dimensions["A"].width = 12

    # -- Remaining datasets -------------------------------------------------
    simple = [
        ("T3 Fuel Consumption", "fuel-consumption.csv",
         "Fuel consumption of a sample of 27 cars",
         "The population standard deviation is given in the question as 1.3 L/100km.",
         ["Vehicle", "Fuel consumption (L/100km)"]),
        ("T3 HGV Weights", "hgv-weights.csv",
         "Weight of 38 randomly selected heavy goods vehicles",
         "Weights in tonnes.",
         ["Vehicle", "Weight (tonnes)"]),
        ("T4 Junction Waiting", "junction-waiting.csv",
         "Waiting time at a junction, before and after a development project",
         "Each row is the same location measured twice - these are paired observations.",
         ["Observation", "Before (s)", "After (s)"]),
        ("T5 Taxi Queue", "taxi-queue.csv",
         "Number of people in a queue at a taxi rank",
         "", ["Number in queue", "Frequency"]),
        ("T5 Car Ownership", "car-ownership.csv",
         "Car ownership by employment status",
         "Row and column totals are for you to complete.",
         ["Status", "Has car", "Does not have car"]),
        ("T5 Accidents Weather", "accidents-weather.csv",
         "Road accidents by road surface condition",
         "", ["City", "Dry", "Wet", "Snow", "Ice"]),
        ("T5 Reaction Distance", "reaction-distance.csv",
         "Warning distance and average reaction distance for motorway closures",
         "", ["Site", "Warning distance (m)", "Average reaction distance (m)"]),
    ]
    for sheet, csv_name, title, note, headers in simple:
        ws = wb.create_sheet(sheet)
        ws["A1"] = title
        ws["A1"].font = TITLE_FONT
        start = 3
        if note:
            ws["A2"] = note
            ws["A2"].font = NOTE_FONT
        for c, h in enumerate(headers, start=1):
            ws.cell(row=start, column=c, value=h)
        style_header(ws, start, len(headers))
        _, rows = read_csv(csv_name)
        for i, row in enumerate(rows):
            for c, val in enumerate(row, start=1):
                ws.cell(row=start + 1 + i, column=c, value=val).border = BOX
        autosize(ws, 22)

    return wb


# ---------------------------------------------------------------------------
# Hypothesis test workbench
# ---------------------------------------------------------------------------

def _sheet_frame(ws, title: str, blurb: str, when: str) -> int:
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = blurb
    ws["A2"].font = NOTE_FONT
    ws["A3"] = f"Use when: {when}"
    ws["A3"].font = NOTE_FONT
    ws["A5"] = "INPUTS"
    ws["A5"].font = BOLD
    return 6


def _inputs(ws, row: int, items: list[tuple[str, object, str]]) -> dict[str, str]:
    """Write label/value/note rows. Returns {label: cell ref}."""
    refs = {}
    for label, value, note in items:
        ws.cell(row=row, column=1, value=label).border = BOX
        c = ws.cell(row=row, column=2, value=value)
        c.fill = INPUT_FILL
        c.border = BOX
        c.font = BOLD
        if note:
            n = ws.cell(row=row, column=3, value=note)
            n.font = NOTE_FONT
        refs[label] = f"B{row}"
        row += 1
    return refs, row


def _outputs(ws, row: int, items: list[tuple[str, str, str]], fill=RESULT_FILL) -> int:
    for label, formula, note in items:
        ws.cell(row=row, column=1, value=label).border = BOX
        c = ws.cell(row=row, column=2, value=formula)
        c.fill = fill
        c.border = BOX
        if note:
            n = ws.cell(row=row, column=3, value=note)
            n.font = NOTE_FONT
        row += 1
    return row


def _verdict(ws, row: int, stat_ref: str, crit_ref: str, p_ref: str,
             alpha_ref: str, two_tailed: bool = True) -> int:
    ws.cell(row=row, column=1, value="VERDICT").font = BOLD
    row += 1
    cmp_stat = f"ABS({stat_ref})" if two_tailed else stat_ref
    ws.cell(row=row, column=1, value="By critical value").border = BOX
    c = ws.cell(row=row, column=2,
                value=f'=IF({cmp_stat}>{crit_ref},"Reject H0","Do not reject H0")')
    c.fill = VERDICT_FILL
    c.border = BOX
    ws.cell(row=row, column=3,
            value="Reject when the test statistic is further from zero than the critical value.").font = NOTE_FONT
    row += 1
    ws.cell(row=row, column=1, value="By p-value").border = BOX
    c = ws.cell(row=row, column=2,
                value=f'=IF({p_ref}<{alpha_ref},"Reject H0","Do not reject H0")')
    c.fill = VERDICT_FILL
    c.border = BOX
    ws.cell(row=row, column=3, value="Reject when p < alpha.").font = NOTE_FONT
    row += 1
    ws.cell(row=row, column=1, value="Routes agree?").border = BOX
    c = ws.cell(row=row, column=2, value=f'=IF(B{row-2}=B{row-1},"yes","CHECK YOUR INPUTS")')
    c.fill = VERDICT_FILL
    c.border = BOX
    ws.cell(row=row, column=3,
            value="These two always agree. If they don't, an input is wrong.").font = NOTE_FONT
    row += 2
    ws.cell(row=row, column=1,
            value="Remember: failing to reject H0 is not the same as proving it true.").font = NOTE_FONT
    return row + 1


def build_workbench() -> None:
    wb = Workbook()

    # -- Start here ---------------------------------------------------------
    ws = wb.active
    ws.title = "Start Here"
    ws["A1"] = "Hypothesis Test Workbench"
    ws["A1"].font = TITLE_FONT
    lines = [
        "",
        "Excel has no function that runs a hypothesis test from summary statistics.",
        "T.TEST needs two ranges of raw data, and so does the Analysis ToolPak. When a",
        "question gives you only n, the mean and the standard deviation, you have to",
        "build the test out of cell formulas. That is what these sheets do.",
        "",
        "How to use a sheet",
        "  1. Pick the sheet matching your test (see the chooser below).",
        "  2. Type your numbers into the yellow cells only.",
        "  3. Read the test statistic, the critical value and the p-value.",
        "  4. Check that both verdicts agree. They always should.",
        "",
        "Which test do I need?",
        "  Comparing one sample mean to a known population mean",
        "      population sd known ................ One-Sample Z",
        "      population sd unknown .............. One-Sample T",
        "  Comparing two independent groups",
        "      population sds known, or n large ... Two-Sample Z",
        "      population sds unknown ............. Two-Sample T",
        "  Same subjects measured twice ........... Paired T",
        "  Two categorical variables .............. Chi-Squared",
        "  Is a regression coefficient real? ...... Regression Coefficient T",
        "",
        "Every sheet reports both the critical-value route and the p-value route.",
        "They are two ways of asking the same question, and a result you can reach",
        "two ways is more trustworthy than one you can only reach one way.",
    ]
    for i, line in enumerate(lines, start=2):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 84

    # -- One-sample z -------------------------------------------------------
    ws = wb.create_sheet("One-Sample Z")
    r = _sheet_frame(ws, "One-Sample Z-Test",
                     "Compares a sample mean against a known population mean.",
                     "the population standard deviation is known.")
    refs, r = _inputs(ws, r, [
        ("Sample mean (xbar)", 10.9, ""),
        ("Population mean (mu0)", 12, "the value under H0"),
        ("Population sd (sigma)", 3.5, "known, not estimated from the sample"),
        ("Sample size (n)", 100, ""),
        ("Alpha", 0.05, "0.05 for 95% confidence"),
        ("Tails (1 or 2)", 1, "1 for a directional H1, 2 otherwise"),
    ])
    xb, mu, sg, n, al, tl = (refs["Sample mean (xbar)"], refs["Population mean (mu0)"],
                             refs["Population sd (sigma)"], refs["Sample size (n)"],
                             refs["Alpha"], refs["Tails (1 or 2)"])
    r += 1
    ws.cell(row=r, column=1, value="RESULTS").font = BOLD
    r += 1
    se, z, crit, p = f"B{r}", f"B{r+1}", f"B{r+2}", f"B{r+3}"
    r = _outputs(ws, r, [
        ("Standard error", f"={sg}/SQRT({n})", "sigma / sqrt(n)"),
        ("Test statistic z", f"=({xb}-{mu})/{se}", "can be negative - that is fine"),
        ("Critical value", f"=IF({tl}=2,NORM.S.INV(1-{al}/2),NORM.S.INV(1-{al}))",
         "two-tailed splits alpha between the tails"),
        ("p-value", f"=IF({tl}=2,2*(1-NORM.S.DIST(ABS({z}),TRUE)),1-NORM.S.DIST(ABS({z}),TRUE))",
         "probability of a result at least this extreme if H0 is true"),
    ])
    r += 1
    _verdict(ws, r, z, crit, p, al)
    autosize(ws, 24)
    ws.column_dimensions["C"].width = 62

    # -- One-sample t -------------------------------------------------------
    ws = wb.create_sheet("One-Sample T")
    r = _sheet_frame(ws, "One-Sample T-Test",
                     "Compares a sample mean against a hypothesised population mean.",
                     "the population standard deviation is unknown and estimated from the sample.")
    refs, r = _inputs(ws, r, [
        ("Sample mean (xbar)", 25.0, ""),
        ("Hypothesised mean (mu0)", 24.0, "the value under H0"),
        ("Sample sd (s)", 2.0, "STDEV.S of your data"),
        ("Sample size (n)", 30, ""),
        ("Alpha", 0.05, ""),
        ("Tails (1 or 2)", 2, ""),
    ])
    xb, mu, s, n, al, tl = (refs["Sample mean (xbar)"], refs["Hypothesised mean (mu0)"],
                            refs["Sample sd (s)"], refs["Sample size (n)"],
                            refs["Alpha"], refs["Tails (1 or 2)"])
    r += 1
    ws.cell(row=r, column=1, value="RESULTS").font = BOLD
    r += 1
    se, df, t, crit, p = f"B{r}", f"B{r+1}", f"B{r+2}", f"B{r+3}", f"B{r+4}"
    r = _outputs(ws, r, [
        ("Standard error", f"={s}/SQRT({n})", "s / sqrt(n)"),
        ("Degrees of freedom", f"={n}-1", "n - 1"),
        ("Test statistic t", f"=({xb}-{mu})/{se}", ""),
        ("Critical value", f"=IF({tl}=2,T.INV.2T({al},{df}),T.INV(1-{al},{df}))",
         "T.INV.2T already halves alpha - do not halve it yourself"),
        ("p-value", f"=IF({tl}=2,T.DIST.2T(ABS({t}),{df}),T.DIST.RT(ABS({t}),{df}))", ""),
    ])
    r += 1
    _verdict(ws, r, t, crit, p, al)
    autosize(ws, 24)
    ws.column_dimensions["C"].width = 62

    # -- Two-sample z -------------------------------------------------------
    ws = wb.create_sheet("Two-Sample Z")
    r = _sheet_frame(ws, "Two-Sample Z-Test",
                     "Compares the means of two independent samples.",
                     "population sds are known, or both samples are large enough for the CLT.")
    refs, r = _inputs(ws, r, [
        ("Mean 1", 3.9, ""), ("SD 1", 2.1, ""), ("n 1", 50, ""),
        ("Mean 2", 4.8, ""), ("SD 2", 1.9, ""), ("n 2", 40, ""),
        ("Alpha", 0.05, ""), ("Tails (1 or 2)", 2, ""),
    ])
    m1, s1, n1 = refs["Mean 1"], refs["SD 1"], refs["n 1"]
    m2, s2, n2 = refs["Mean 2"], refs["SD 2"], refs["n 2"]
    al, tl = refs["Alpha"], refs["Tails (1 or 2)"]
    r += 1
    ws.cell(row=r, column=1, value="RESULTS").font = BOLD
    r += 1
    se, z, crit, p = f"B{r}", f"B{r+1}", f"B{r+2}", f"B{r+3}"
    r = _outputs(ws, r, [
        ("Standard error of difference", f"=SQRT({s1}^2/{n1}+{s2}^2/{n2})",
         "sqrt(s1^2/n1 + s2^2/n2)"),
        ("Test statistic z", f"=({m1}-{m2})/{se}", ""),
        ("Critical value", f"=IF({tl}=2,NORM.S.INV(1-{al}/2),NORM.S.INV(1-{al}))",
         "1.96 for a two-tailed test at alpha = 0.05"),
        ("p-value", f"=IF({tl}=2,2*(1-NORM.S.DIST(ABS({z}),TRUE)),1-NORM.S.DIST(ABS({z}),TRUE))", ""),
    ])
    r += 1
    r = _verdict(ws, r, z, crit, p, al)
    ws.cell(row=r + 1, column=1,
            value=("Note: with unknown population sds a t-test is the more standard choice. "
                   "At these sample sizes the two give nearly the same answer.")).font = NOTE_FONT
    autosize(ws, 26)
    ws.column_dimensions["C"].width = 62

    # -- Two-sample t (pooled) ---------------------------------------------
    ws = wb.create_sheet("Two-Sample T")
    r = _sheet_frame(ws, "Two-Sample T-Test (pooled variance)",
                     "Compares the means of two independent samples.",
                     "population sds are unknown but the two groups have similar variance.")
    refs, r = _inputs(ws, r, [
        ("Mean 1", 144, ""), ("SD 1", 49, ""), ("n 1", 64, ""),
        ("Mean 2", 121, ""), ("SD 2", 49, ""), ("n 2", 49, ""),
        ("Alpha", 0.05, ""), ("Tails (1 or 2)", 2, ""),
    ])
    m1, s1, n1 = refs["Mean 1"], refs["SD 1"], refs["n 1"]
    m2, s2, n2 = refs["Mean 2"], refs["SD 2"], refs["n 2"]
    al, tl = refs["Alpha"], refs["Tails (1 or 2)"]
    r += 1
    ws.cell(row=r, column=1, value="RESULTS").font = BOLD
    r += 1
    sp, se, df, t, crit, p = (f"B{r}", f"B{r+1}", f"B{r+2}",
                              f"B{r+3}", f"B{r+4}", f"B{r+5}")
    r = _outputs(ws, r, [
        ("Pooled sd", f"=SQRT((({n1}-1)*{s1}^2+({n2}-1)*{s2}^2)/({n1}+{n2}-2))",
         "when both sds are equal this is just that value"),
        ("Standard error of difference", f"={sp}*SQRT(1/{n1}+1/{n2})", ""),
        ("Degrees of freedom", f"={n1}+{n2}-2", "n1 + n2 - 2"),
        ("Test statistic t", f"=({m1}-{m2})/{se}", ""),
        ("Critical value", f"=IF({tl}=2,T.INV.2T({al},{df}),T.INV(1-{al},{df}))", ""),
        ("p-value", f"=IF({tl}=2,T.DIST.2T(ABS({t}),{df}),T.DIST.RT(ABS({t}),{df}))", ""),
    ])
    r += 1
    r = _verdict(ws, r, t, crit, p, al)
    ws.cell(row=r + 1, column=1,
            value="Preloaded with the Leeds/Manchester bus example: t = 2.473, critical 1.98.").font = NOTE_FONT
    autosize(ws, 26)
    ws.column_dimensions["C"].width = 62

    # -- Paired t -----------------------------------------------------------
    ws = wb.create_sheet("Paired T")
    r = _sheet_frame(ws, "Paired-Sample T-Test",
                     "Tests whether the mean of the paired differences is zero.",
                     "the same subjects are measured twice - before and after.")
    ws.cell(row=r, column=1,
            value="Work out the difference for each pair first, then summarise those differences.").font = NOTE_FONT
    r += 1
    refs, r = _inputs(ws, r, [
        ("Mean of differences", -8.4, "AVERAGE of the difference column"),
        ("SD of differences", 4.99, "STDEV.S of the difference column"),
        ("Number of pairs (n)", 30, "pairs, not total observations"),
        ("Alpha", 0.05, ""),
        ("Tails (1 or 2)", 2, ""),
    ])
    md, sd_, n = (refs["Mean of differences"], refs["SD of differences"],
                  refs["Number of pairs (n)"])
    al, tl = refs["Alpha"], refs["Tails (1 or 2)"]
    r += 1
    ws.cell(row=r, column=1, value="RESULTS").font = BOLD
    r += 1
    se, df, t, crit, p = f"B{r}", f"B{r+1}", f"B{r+2}", f"B{r+3}", f"B{r+4}"
    r = _outputs(ws, r, [
        ("Standard error", f"={sd_}/SQRT({n})", ""),
        ("Degrees of freedom", f"={n}-1", "n - 1, where n is the number of pairs"),
        ("Test statistic t", f"={md}/{se}", ""),
        ("Critical value", f"=IF({tl}=2,T.INV.2T({al},{df}),T.INV(1-{al},{df}))", ""),
        ("p-value", f"=IF({tl}=2,T.DIST.2T(ABS({t}),{df}),T.DIST.RT(ABS({t}),{df}))", ""),
    ])
    r += 1
    r = _verdict(ws, r, t, crit, p, al)
    ws.cell(row=r + 1, column=1,
            value=("If you have both raw columns you can cross-check with "
                   "=T.TEST(range1, range2, 2, 1) - the 1 means paired.")).font = NOTE_FONT
    autosize(ws, 26)
    ws.column_dimensions["C"].width = 62

    # -- Chi-squared --------------------------------------------------------
    ws = wb.create_sheet("Chi-Squared")
    r = _sheet_frame(ws, "Chi-Squared Test for Independence",
                     "Tests whether two categorical variables are associated.",
                     "you have counts in a contingency table and every expected value is at least 5.")
    refs, r = _inputs(ws, r, [
        ("Test statistic (sum of (O-E)^2/E)", 4.1, "compute this in your own table"),
        ("Number of rows", 2, "after any pooling"),
        ("Number of columns", 3, "after any pooling"),
        ("Alpha", 0.05, ""),
    ])
    stat, nr, nc, al = (refs["Test statistic (sum of (O-E)^2/E)"], refs["Number of rows"],
                        refs["Number of columns"], refs["Alpha"])
    r += 1
    ws.cell(row=r, column=1, value="RESULTS").font = BOLD
    r += 1
    df, crit, p = f"B{r}", f"B{r+1}", f"B{r+2}"
    r = _outputs(ws, r, [
        ("Degrees of freedom", f"=({nr}-1)*({nc}-1)", "(rows - 1) x (columns - 1)"),
        ("Critical value", f"=CHISQ.INV.RT({al},{df})",
         "CHISQ.INV.RT, not CHI.INV.RT - the latter does not exist"),
        ("p-value", f"=CHISQ.DIST.RT({stat},{df})", ""),
    ])
    r += 1
    r = _verdict(ws, r, stat, crit, p, al, two_tailed=False)
    for line in [
        "Chi-squared is never two-tailed: the statistic cannot be negative, so there is no left tail.",
        "If any expected value is below 5, pool similar categories together and recount the",
        "degrees of freedom before using this sheet.",
    ]:
        r += 1
        ws.cell(row=r, column=1, value=line).font = NOTE_FONT
    autosize(ws, 30)
    ws.column_dimensions["C"].width = 62

    # -- Regression coefficient t ------------------------------------------
    ws = wb.create_sheet("Regression Coeff T")
    r = _sheet_frame(ws, "Regression Coefficient T-Test",
                     "Tests whether a fitted coefficient is really different from zero.",
                     "you have fitted r = a*w + b and want to know if the slope is real.")
    refs, r = _inputs(ws, r, [
        ("Coefficient", 0.164, "the gradient a, or the intercept b"),
        ("Sum of squared residuals", 538.6, "sum of e^2"),
        ("S_xx", 107702.5, "sum(x^2) - n * xbar^2"),
        ("Sample size (n)", 10, ""),
        ("Alpha", 0.05, ""),
    ])
    coef, rss, sxx, n, al = (refs["Coefficient"], refs["Sum of squared residuals"],
                             refs["S_xx"], refs["Sample size (n)"], refs["Alpha"])
    r += 1
    ws.cell(row=r, column=1, value="RESULTS").font = BOLD
    r += 1
    s, se, df, t, crit, p = (f"B{r}", f"B{r+1}", f"B{r+2}",
                             f"B{r+3}", f"B{r+4}", f"B{r+5}")
    r = _outputs(ws, r, [
        ("Residual standard error", f"=SQRT({rss}/({n}-2))", "sqrt(RSS / (n-2))"),
        ("Standard error of coefficient", f"={s}/SQRT({sxx})", ""),
        ("Degrees of freedom", f"={n}-2", "n - 2: two coefficients were estimated"),
        ("Test statistic t", f"={coef}/{se}", "coefficient / its standard error"),
        ("Critical value", f"=T.INV.2T({al},{df})", "always two-tailed here"),
        ("p-value", f"=T.DIST.2T(ABS({t}),{df})", ""),
    ])
    r += 1
    r = _verdict(ws, r, t, crit, p, al)
    ws.cell(row=r + 1, column=1,
            value="Preloaded with the Tutorial 5 warning-distance regression: t = 6.56, critical 2.306.").font = NOTE_FONT
    autosize(ws, 30)
    ws.column_dimensions["C"].width = 62

    return wb


# ---------------------------------------------------------------------------

TARGETS = [
    (build_data_workbook, XLSX_DIR / "transport-stats-data.xlsx"),
    (build_workbench, WB_DIR / "hypothesis-test-workbench.xlsx"),
]


def cell_contents(wb) -> dict:
    """Every non-blank cell value, keyed by sheet and coordinate.

    Empty strings count as blank: openpyxl writes '' but reads it back as
    None, so keeping them would make every workbook look changed.
    """
    return {
        (ws.title, cell.coordinate): cell.value
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None and cell.value != ""
    }


def check() -> int:
    """Confirm the committed workbooks still match what the code would build.

    Compares cell *contents*, not bytes: an .xlsx is a ZIP archive carrying
    timestamps, so two builds of identical content are never byte-identical.
    """
    from openpyxl import load_workbook

    print("Checking the committed workbooks against a fresh build...")
    problems = []
    for builder, path in TARGETS:
        if not path.exists():
            problems.append(f"{path.relative_to(ROOT)} is missing")
            continue
        fresh = cell_contents(builder())
        committed = cell_contents(load_workbook(path))
        if fresh == committed:
            print(f"  {path.relative_to(ROOT)}: {len(fresh)} cells match")
            continue
        only_fresh = set(fresh) - set(committed)
        only_committed = set(committed) - set(fresh)
        changed = [k for k in set(fresh) & set(committed) if fresh[k] != committed[k]]
        problems.append(
            f"{path.relative_to(ROOT)}: {len(changed)} cells differ, "
            f"{len(only_fresh)} added, {len(only_committed)} removed")
        for key in (changed + sorted(only_fresh) + sorted(only_committed))[:8]:
            problems.append(f"    {key[0]}!{key[1]}: "
                            f"committed {committed.get(key)!r} -> fresh {fresh.get(key)!r}")

    if problems:
        print("\nThe committed workbooks are out of date:\n")
        for p in problems:
            print(f"  x {p}")
        print("\nRun 'python tools/build_workbooks.py' and commit the result.")
        return 1
    print("Up to date.")
    return 0


def main() -> int:
    if "--check" in sys.argv:
        return check()
    print("Building workbooks...")
    for builder, path in TARGETS:
        path.parent.mkdir(parents=True, exist_ok=True)
        wb = builder()
        wb.save(path)
        print(f"  wrote {path.relative_to(ROOT)}  ({len(wb.sheetnames)} sheets)")
    print("Done.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
