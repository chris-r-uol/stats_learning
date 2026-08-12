"""
Prove that every answer the site tells a student is correct.

Three layers:

  1. DATASETS      -- the generated data reproduces every summary statistic the
                      original document publishes.
  2. ANSWER KEYS   -- every standalone numeric answer across the five tutorials
                      is recomputed from first principles with numpy/scipy.
  3. CONTENT       -- every `verify:` expression in the YAML question files is
                      evaluated and checked against the `answer:` stored next to
                      it, plus structural schema checks.

Also lints the Excel formulas in the workbooks against a list of real function
names, because a plausible-looking function that does not exist (CHI.INV.RT)
is exactly the kind of error this project is meant to remove.

Run:  python tools/verify_answers.py
Exit code is non-zero if anything fails, so CI can gate on it.
"""

from __future__ import annotations

import csv
import math
import re
import sys
from pathlib import Path

import numpy as np
from scipy import stats

ROOT = Path(__file__).resolve().parents[1]
CSV_DIR = ROOT / "data" / "csv"
CONTENT_DIR = ROOT / "content"

FAILURES: list[str] = []
CHECKS = 0


def check(label: str, got, want, tol=5e-3) -> None:
    """Assert got ~= want. Tolerance is absolute and defaults to half a unit in
    the second decimal place, matching how the document quotes its answers."""
    global CHECKS
    CHECKS += 1
    if want is None:
        return
    try:
        ok = abs(float(got) - float(want)) <= tol
    except (TypeError, ValueError):
        ok = got == want
    if not ok:
        FAILURES.append(f"{label}: got {got!r}, expected {want!r} (tol {tol})")


def section(name: str) -> None:
    print(f"\n{name}")
    print("-" * len(name))


# ---------------------------------------------------------------------------
# Excel-compatible statistics
# ---------------------------------------------------------------------------

def stdev_s(x):
    return float(np.std(np.asarray(x, dtype=float), ddof=1))


def var_s(x):
    return float(np.var(np.asarray(x, dtype=float), ddof=1))


def skew_excel(x):
    x = np.asarray(x, dtype=float)
    n = len(x)
    z = (x - x.mean()) / stdev_s(x)
    return float(n / ((n - 1) * (n - 2)) * np.sum(z**3))


def quartile_inc(x, q):
    return float(np.percentile(np.sort(np.asarray(x, dtype=float)), q * 100,
                               method="linear"))


def mode_sngl(x):
    vals, counts = np.unique(np.asarray(x), return_counts=True)
    return float(vals[counts.argmax()])


# Excel function equivalents, named so the YAML `verify:` expressions read like
# the formulas the students are typing.
NORM_DIST = lambda x, m, s, cum=True: float(stats.norm.cdf(x, m, s) if cum else stats.norm.pdf(x, m, s))
NORM_INV = lambda p, m, s: float(stats.norm.ppf(p, m, s))
NORM_S_DIST = lambda z, cum=True: float(stats.norm.cdf(z) if cum else stats.norm.pdf(z))
NORM_S_INV = lambda p: float(stats.norm.ppf(p))
BINOM_DIST = lambda k, n, p, cum: float(stats.binom.cdf(k, n, p) if cum else stats.binom.pmf(k, n, p))
POISSON_DIST = lambda k, lam, cum: float(stats.poisson.cdf(k, lam) if cum else stats.poisson.pmf(k, lam))
T_INV = lambda p, df: float(stats.t.ppf(p, df))
T_INV_2T = lambda a, df: float(stats.t.ppf(1 - a / 2, df))
T_DIST_2T = lambda t, df: float(2 * stats.t.sf(abs(t), df))
T_DIST_RT = lambda t, df: float(stats.t.sf(t, df))
CHISQ_INV_RT = lambda a, df: float(stats.chi2.isf(a, df))
CHISQ_DIST_RT = lambda x, df: float(stats.chi2.sf(x, df))


def load(name: str):
    with open(CSV_DIR / name) as fh:
        rows = list(csv.reader(fh))
    header, body = rows[0], rows[1:]
    cols = {h: [] for h in header}
    for r in body:
        for h, v in zip(header, r):
            try:
                cols[h].append(float(v))
            except ValueError:
                cols[h].append(v)
    return cols


# ---------------------------------------------------------------------------
# Layer 1 -- datasets reproduce the published summary statistics
# ---------------------------------------------------------------------------

def verify_datasets(ns: dict) -> None:
    section("Datasets reproduce the published statistics")

    travel = ns["travel"]
    check("T1 n", len(travel), 55, 0)
    check("T1 mean", travel.mean(), 22.58)
    check("T1 median", np.median(travel), 19)
    check("T1 mode", mode_sngl(travel), 12, 0)
    check("T1 variance", var_s(travel), 113.99)
    # The document prints 10.67, but sqrt(113.99) = 10.6766 -> 10.68.
    # See ERRATA.md item 6; we accept the range that spans both.
    check("T1 sd", stdev_s(travel), 10.675, 0.01)
    check("T1 skew", skew_excel(travel), 1.20)
    check("T1 Q1", quartile_inc(travel, 0.25), 14.5)
    check("T1 Q2", quartile_inc(travel, 0.50), 19)
    check("T1 Q3", quartile_inc(travel, 0.75), 30.5)
    check("T1 max is the outlier", travel.max(), 61, 0)

    no_out = ns["travel_no_outlier"]
    check("T1 no-outlier n", len(no_out), 54, 0)
    check("T1 no-outlier mean", no_out.mean(), 21.87)
    check("T1 no-outlier median", np.median(no_out), 18.50)
    check("T1 no-outlier sd", stdev_s(no_out), 9.37)
    check("T1 no-outlier variance", var_s(no_out), 87.78)
    check("T1 no-outlier skew", skew_excel(no_out), 0.69, 0.01)
    check("T1 no-outlier Q1", quartile_inc(no_out, 0.25), 14.25)
    check("T1 no-outlier Q3", quartile_inc(no_out, 0.75), 30.00)

    # Histogram: relative frequency, density and cumulative (Q3/Q4)
    counts = np.array([0, 0, 16, 16, 4, 5, 7, 4, 2, 0, 0, 0, 1], dtype=float)
    # The document's bins are labelled (lo, hi] -- left-open, right-closed --
    # which is the opposite convention to numpy's default, so bin them by hand.
    edges = np.arange(0, 70, 5)
    binned = np.array([np.sum((travel > lo) & (travel <= hi))
                       for lo, hi in zip(edges[:-1], edges[1:])], dtype=float)
    check("T1 histogram counts match", int(np.abs(binned - counts).sum()), 0, 0)
    relfreq = counts / counts.sum()
    published = [0.000, 0.000, 0.291, 0.291, 0.073, 0.091, 0.127, 0.073,
                 0.036, 0.000, 0.000, 0.000, 0.018]
    for i, want in enumerate(published):
        check(f"T1 relative frequency bin {i}", relfreq[i], want, 1e-3)
        check(f"T1 relative frequency density bin {i}", relfreq[i] / 5, want / 5, 1e-3)
    cum = np.cumsum(relfreq)
    for i, want in enumerate([0.000, 0.000, 0.291, 0.582, 0.655, 0.745, 0.873,
                              0.945, 0.982, 0.982, 0.982, 0.982, 1.000]):
        check(f"T1 cumulative bin {i}", cum[i], want, 1.5e-3)

    fuel = ns["fuel"]
    check("T3 fuel n", len(fuel), 27, 0)
    check("T3 fuel mean", fuel.mean(), 10.08)

    hgv = ns["hgv"]
    check("T3 HGV n", len(hgv), 38, 0)
    check("T3 HGV mean", hgv.mean(), 22.29)
    check("T3 HGV sd", stdev_s(hgv), 7.59)

    diff = ns["junction_diff"]
    check("T4 junction n", len(diff), 30, 0)
    check("T4 junction mean difference", diff.mean(), -8.4)
    check("T4 junction sd of differences", stdev_s(diff), 4.99)

    print(f"  {CHECKS} checks so far")


# ---------------------------------------------------------------------------
# Layer 2 -- standalone answer keys, tutorial by tutorial
# ---------------------------------------------------------------------------

def verify_tutorial_2() -> None:
    section("Tutorial 2 -- distributions")
    check("T2 Q2a z-score", (15 - 25) / 12, -0.833, 5e-3)
    check("T2 Q2b P(X<15)", NORM_DIST(15, 25, 12), 0.2023, 1e-3)
    check("T2 Q2b complement", 1 - NORM_DIST(15, 25, 12), 0.798, 1e-3)
    # The document prints 44.73; the true value is 44.7382, which rounds to
    # 44.74. See ERRATA.md item 13 -- the site accepts both.
    check("T2 Q2c NORM.INV(0.95)", NORM_INV(0.95, 25, 12), 44.738, 5e-3)

    check("T2 Q3a total combinations", math.log10(2**85), math.log10(3.87e25), 2e-3)
    # The document walks through the formula for this part but never states the
    # number, so a student cannot check themselves. ERRATA.md item 14.
    check("T2 Q3b P(X<10)", BINOM_DIST(9, 85, 0.14, True), 0.2316, 5e-4)
    check("T2 Q3c P(X<=15)", BINOM_DIST(15, 85, 0.14, True), 0.868, 2e-3)
    check("T2 Q3c P(X>15)", 1 - BINOM_DIST(15, 85, 0.14, True), 0.132, 2e-3)
    check("T2 Q3d P(X=8)", BINOM_DIST(8, 85, 0.14, False), 0.064235, 1e-4)

    check("T2 Q4a P(X=1|3)", POISSON_DIST(1, 3, False), 0.1494, 1e-3)
    check("T2 Q4b P(X>10|7)", 1 - POISSON_DIST(10, 7, True), 0.0985, 2e-3)
    check("T2 Q4c P(X<=7|4)", POISSON_DIST(7, 4, True), 0.9489, 1e-3)
    check("T2 Q4c P(X<=8|4)", POISSON_DIST(8, 4, True), 0.9786, 1e-3)


def verify_tutorial_3(ns: dict) -> None:
    section("Tutorial 3 -- interval estimation")
    check("T3 Q1 lower", NORM_INV(0.025, 25, 3), 19.12, 5e-3)
    check("T3 Q1 upper", NORM_INV(0.975, 25, 3), 30.88, 5e-3)
    check("T3 Q1 c", NORM_INV(0.975, 25, 3) - 25, 5.88, 5e-3)

    check("T3 Q2 lower", NORM_INV(0.005, -2, 0.1), -2.258, 5e-3)
    check("T3 Q2 upper", NORM_INV(0.995, -2, 0.1), -1.742, 5e-3)
    check("T3 Q2 c", NORM_INV(0.995, -2, 0.1) + 2, 0.258, 5e-3)

    fuel = ns["fuel"]
    se = 1.3 / math.sqrt(27)
    check("T3 Q3 sample mean", fuel.mean(), 10.08)
    check("T3 Q3 standard error", se, 0.25, 5e-3)
    check("T3 Q3 CI lower", NORM_INV(0.025, 10.08, 0.25), 9.59, 5e-3)
    check("T3 Q3 CI upper", NORM_INV(0.975, 10.08, 0.25), 10.57, 5e-3)

    hgv = ns["hgv"]
    se4 = stdev_s(hgv) / math.sqrt(38)
    check("T3 Q4 sample mean", hgv.mean(), 22.29)
    check("T3 Q4 standard error", se4, 1.23, 5e-3)
    check("T3 Q4 CI lower", NORM_INV(0.025, 22.29, 1.23), 19.88, 5e-3)
    check("T3 Q4 CI upper", NORM_INV(0.975, 22.29, 1.23), 24.70, 5e-3)

    se5 = 8 / math.sqrt(38)
    check("T3 Q5 standard error", se5, 1.29777, 1e-4)
    check("T3 Q5 CI lower", NORM_INV(0.025, 22.29, se5), 19.75, 5e-3)
    check("T3 Q5 CI upper", NORM_INV(0.975, 22.29, se5), 24.83, 5e-3)

    # ERRATA item 16: the worked example in the methodology section builds its
    # interval from the standard deviation (2) rather than the standard error
    # (2/sqrt(100) = 0.2), making it sqrt(n) = 10 times too wide.
    check("Methodology CI standard error", 2 / math.sqrt(100), 0.2, 1e-12)
    check("Methodology CI lower (correct)", NORM_INV(0.025, 25, 0.2), 24.61, 5e-3)
    check("Methodology CI upper (correct)", NORM_INV(0.975, 25, 0.2), 25.39, 5e-3)
    check("Methodology CI half-width (correct)", NORM_INV(0.975, 25, 0.2) - 25,
          0.392, 5e-3)
    check("Methodology CI as printed in the handout",
          NORM_INV(0.975, 25, 2) - 25, 3.92, 5e-3)
    check("Handout interval is sqrt(n) too wide",
          (NORM_INV(0.975, 25, 2) - 25) / (NORM_INV(0.975, 25, 0.2) - 25), 10, 1e-6)

    check("T3 Q6 sd from variance", math.sqrt(9), 3, 0)
    for k, halfwidth in [(1, 3), (2, 6), (3, 9), (4, 12)]:
        check(f"T3 Q6 {k}-sigma half width", k * 3, halfwidth, 0)


def verify_tutorial_4(ns: dict) -> None:
    section("Tutorial 4 -- hypothesis testing")
    # Worked example in the methodology section: Leeds vs Manchester buses
    t = (144 - 121) / (49 * math.sqrt(1 / 64 + 1 / 49))
    check("Methodology example t", t, 2.473, 5e-3)
    check("Methodology example df", 64 + 49 - 2, 111, 0)
    check("Methodology example critical t", T_INV_2T(0.05, 111), 1.98, 5e-3)

    # Q3: one-sample z, left tailed
    z = (10.9 - 12) / (3.5 / math.sqrt(100))
    check("T4 Q3 z", z, -3.14, 5e-3)
    check("T4 Q3 p", NORM_S_DIST(-3.14), 0.0008, 5e-5)
    check("T4 Q3 p exact", NORM_S_DIST(z), 0.000835, 5e-5)

    # Q4: two-sample z, two tailed
    se = math.sqrt(2.1**2 / 50 + 1.9**2 / 40)
    z4 = (4.8 - 3.9) / se
    check("T4 Q4 standard error", se, 0.422, 5e-3)
    check("T4 Q4 z", z4, 2.13, 5e-3)
    check("T4 Q4 critical z", NORM_S_INV(0.975), 1.96, 5e-3)
    check("T4 Q4 p", 2 * (1 - NORM_S_DIST(z4)), 0.033, 1e-3)

    # Q5: paired t on the real data
    d = ns["junction_diff"]
    t5 = d.mean() / (stdev_s(d) / math.sqrt(len(d)))
    check("T4 Q5 mean difference", d.mean(), -8.4)
    check("T4 Q5 sd of differences", stdev_s(d), 4.99)
    check("T4 Q5 t", t5, -9.23, 0.02)
    check("T4 Q5 df", len(d) - 1, 29, 0)
    check("T4 Q5 critical t", T_INV_2T(0.05, 29), 2.045, 5e-3)
    p5 = T_DIST_2T(t5, len(d) - 1)
    check("T4 Q5 p-value magnitude", math.log10(p5), math.log10(4e-10), 0.35)
    # cross-check against scipy's paired test on the raw columns, which is what
    # Excel's T.TEST(range1, range2, 2, 1) computes
    tt = stats.ttest_rel(ns["junction_after"], ns["junction_before"])
    check("T4 Q5 t matches T.TEST route", tt.statistic, t5, 1e-9)
    check("T4 Q5 p matches T.TEST route", tt.pvalue, p5, 1e-12)


def verify_tutorial_5(ns: dict) -> None:
    section("Tutorial 5 -- relationships in data")

    # Q1 -- taxi queue relative frequencies
    freq = np.array([6, 4, 5, 2, 2, 1], dtype=float)
    rel = freq / freq.sum()
    check("T5 Q1 N", freq.sum(), 20, 0)
    for i, want in enumerate([0.3, 0.2, 0.25, 0.1, 0.1, 0.05]):
        check(f"T5 Q1 relative frequency {i+1}", rel[i], want, 1e-9)
    check("T5 Q1 P(n<4)", rel[:3].sum(), 0.75, 1e-9)
    check("T5 Q1 P(n>3)", rel[3:].sum(), 0.25, 1e-9)
    check("T5 Q1 P(A=3)P(B=5)", rel[2] * rel[4], 0.025, 1e-9)

    # Q2 -- car ownership contingency table
    obs = np.array([[130, 21], [21, 34], [54, 50], [31, 120], [31, 53], [11, 13]],
                   dtype=float)
    rows = obs.sum(axis=1)
    cols = obs.sum(axis=0)
    total = obs.sum()
    check("T5 Q2 row total employed", rows[0], 151, 0)
    check("T5 Q2 column total has car", cols[0], 278, 0)
    check("T5 Q2 column total no car", cols[1], 291, 0)
    check("T5 Q2 grand total", total, 569, 0)
    check("T5 Q2a proportion students", rows[3] / total, 0.265, 1e-3)
    check("T5 Q2b proportion with car", cols[0] / total, 0.489, 1e-3)
    check("T5 Q2c students with car", obs[3, 0] / rows[3], 0.205, 1e-3)
    # ERRATA item 5: the question asks for the proportion OF CARLESS PEOPLE who
    # are employed, so the denominator is the carless column total, not the
    # employed row total. The document divides by 151 and gets 0.139.
    check("T5 Q2d carless who are employed", obs[0, 1] / cols[1], 0.072, 1e-3)
    check("T5 Q2d document's value (wrong denominator)", obs[0, 1] / rows[0], 0.139, 1e-3)

    exp = np.outer(rows, cols) / total
    for i, want in enumerate([73.775, 26.872, 50.812, 73.775, 41.04, 11.726]):
        check(f"T5 Q2 expected has-car row {i}", exp[i, 0], want, 5e-3)
    # ERRATA item 15: the document prints 53.118 for the homemaker/no-car cell,
    # but 104 * 291 / 569 = 53.188 -- a transposed digit.
    for i, want in enumerate([77.225, 28.128, 53.188, 77.225, 42.960, 12.274]):
        check(f"T5 Q2 expected no-car row {i}", exp[i, 1], want, 5e-3)
    chi_cells = (obs - exp) ** 2 / exp
    check("T5 Q2 chi2 cell employed/car", chi_cells[0, 0], 42.85, 0.02)
    check("T5 Q2 chi2 statistic", chi_cells.sum(), 140.07, 0.02)
    check("T5 Q2 df", (6 - 1) * (2 - 1), 5, 0)
    check("T5 Q2 critical value at 1%", CHISQ_INV_RT(0.01, 5), 15.09, 5e-3)
    check("T5 Q2 min expected exceeds 5", exp.min() > 5, True)

    # Q3 -- accidents by weather, with pooling
    obs3 = np.array([[115, 174, 32, 6], [68, 71, 17, 7]], dtype=float)
    r3, c3, t3 = obs3.sum(axis=1), obs3.sum(axis=0), obs3.sum()
    check("T5 Q3 Leeds total", r3[0], 327, 0)
    check("T5 Q3 Manchester total", r3[1], 163, 0)
    check("T5 Q3 grand total", t3, 490, 0)
    for i, want in enumerate([183, 245, 49, 13]):
        check(f"T5 Q3 column total {i}", c3[i], want, 0)
    exp3 = np.outer(r3, c3) / t3
    for (i, j), want in zip([(0, 0), (0, 1), (0, 2), (0, 3), (1, 0), (1, 1), (1, 2), (1, 3)],
                            [122.1, 163.5, 32.7, 8.7, 60.9, 81.5, 16.3, 4.3]):
        check(f"T5 Q3 expected [{i},{j}]", exp3[i, j], want, 0.05)
    check("T5 Q3 an expected value is below 5 (forces pooling)", exp3.min() < 5, True)

    # pool snow and ice
    obsP = np.array([[115, 174, 38], [68, 71, 24]], dtype=float)
    rP, cP, tP = obsP.sum(axis=1), obsP.sum(axis=0), obsP.sum()
    expP = np.outer(rP, cP) / tP
    check("T5 Q3 pooled expected Leeds snow+ice", expP[0, 2], 41.38, 5e-3)
    check("T5 Q3 pooled expected Manchester snow+ice", expP[1, 2], 20.62, 5e-3)
    check("T5 Q3 pooled min expected exceeds 5", expP.min() > 5, True)
    chiP = (obsP - expP) ** 2 / expP
    check("T5 Q3 chi2 statistic", chiP.sum(), 4.1, 0.02)
    check("T5 Q3 df after pooling", (2 - 1) * (3 - 1), 2, 0)
    check("T5 Q3 critical value at 5%", CHISQ_INV_RT(0.05, 2), 5.99, 5e-3)
    check("T5 Q3 do not reject", chiP.sum() < CHISQ_INV_RT(0.05, 2), True)

    # Q4 -- least squares regression
    w = ns["warning"]
    r = ns["reaction"]
    n = len(w)
    check("T5 Q4 n", n, 10, 0)
    check("T5 Q4 sum w", w.sum(), 1935, 0)
    check("T5 Q4 sum r", r.sum(), 679.11, 5e-3)
    check("T5 Q4 sum w^2", (w**2).sum(), 482125, 0)
    check("T5 Q4 sum w*r", (w * r).sum(), 149092.9, 0.05)
    check("T5 Q4 xbar", w.mean(), 193.5, 1e-9)
    check("T5 Q4 ybar", r.mean(), 67.911, 5e-4)

    s_xx = float((w**2).sum() - n * w.mean() ** 2)
    s_yy = float((r**2).sum() - n * r.mean() ** 2)
    s_xy = float((w * r).sum() - n * w.mean() * r.mean())
    check("T5 Q4 S_xx", s_xx, 107702.5, 0.5)
    check("T5 Q4 S_yy", s_yy, 3442.51, 0.5)
    check("T5 Q4 S_xy", s_xy, 17685.1, 0.5)

    # ERRATA item 1: S_xy / S_xx is the GRADIENT, not the intercept.
    a = s_xy / s_xx
    b = r.mean() - a * w.mean()
    check("T5 Q4 gradient a", a, 0.164, 5e-4)
    check("T5 Q4 intercept b", b, 36.138, 5e-3)
    slope, intercept, rval, _, stderr = stats.linregress(w, r)
    check("T5 Q4 gradient matches LINEST", slope, a, 1e-9)
    check("T5 Q4 intercept matches LINEST", intercept, b, 1e-9)

    pred = a * w + b
    resid = r - pred
    sse = float((resid**2).sum())
    check("T5 Q4 SSE", sse, 538.56, 0.5)
    check("T5 Q4 SST", s_yy, 3442.51, 0.5)
    check("T5 Q4 R2", 1 - sse / s_yy, 0.84, 5e-3)
    check("T5 Q4 R2 matches r^2 from LINEST", rval**2, 1 - sse / s_yy, 1e-9)

    # The document predicts using the coefficient rounded to 3dp (0.164), which
    # is what a student following the worked steps will actually have in their
    # spreadsheet. Both routes are checked: the gap between them grows with w
    # and is worth showing students as a lesson about rounding too early.
    for wi, want in [(90, 50.898), (150, 60.738), (500, 118.138), (1200, 232.938)]:
        check(f"T5 Q4 prediction at w={wi} (rounded a)", 0.164 * wi + b, want, 5e-3)
    for wi, want in [(90, 50.916), (150, 60.768), (500, 118.241), (1200, 233.182)]:
        check(f"T5 Q4 prediction at w={wi} (full precision)", a * wi + b, want, 5e-3)
    check("T5 Q4 rounding gap at w=1200", abs((a - 0.164) * 1200), 0.244, 0.01)

    s_resid = math.sqrt(sse / (n - 2))
    se_a = s_resid / math.sqrt(s_xx)
    check("T5 Q4 residual standard error", s_resid, 8.2, 0.02)
    check("T5 Q4 SE(a)", se_a, 0.0250, 5e-4)
    check("T5 Q4 SE(a) matches LINEST", stderr, se_a, 1e-9)
    check("T5 Q4 t for gradient", a / se_a, 6.56, 0.02)
    check("T5 Q4 df", n - 2, 8, 0)
    check("T5 Q4 critical t", T_INV(0.975, 8), 2.306, 5e-3)
    check("T5 Q4 reject H0 for gradient", a / se_a > T_INV(0.975, 8), True)


# ---------------------------------------------------------------------------
# Workbook formula lint
# ---------------------------------------------------------------------------

REAL_EXCEL_FUNCTIONS = {
    "ABS", "AVERAGE", "BINOM.DIST", "CHISQ.DIST", "CHISQ.DIST.RT", "CHISQ.INV",
    "CHISQ.INV.RT", "CHISQ.TEST", "COMBIN", "CONFIDENCE.NORM", "CONFIDENCE.T",
    "COUNT", "F.TEST", "IF", "INDIRECT", "INTERCEPT", "LINEST", "MAX", "MEDIAN",
    "MIN", "MODE.SNGL", "NORM.DIST", "NORM.INV", "NORM.S.DIST", "NORM.S.INV",
    "POISSON.DIST", "QUARTILE.INC", "ROW", "RSQ", "SKEW", "SKEW.P", "SLOPE",
    "SQRT", "STDEV.P", "STDEV.S", "STEYX", "SUM", "SUMPRODUCT", "T.DIST",
    "T.DIST.2T", "T.DIST.RT", "T.INV", "T.INV.2T", "T.TEST", "VAR.P", "VAR.S",
    "Z.TEST",
}

FUNC_RE = re.compile(r"\b([A-Z][A-Z0-9.]*)\s*\(")


def lint_excel_formulas() -> None:
    section("Excel formula lint")
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  openpyxl not installed, skipping")
        return

    paths = list((ROOT / "workbooks").glob("*.xlsx")) + \
            list((ROOT / "data" / "xlsx").glob("*.xlsx"))
    seen = set()
    for path in paths:
        wb = load_workbook(path)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        for fn in FUNC_RE.findall(cell.value):
                            seen.add(fn)
                            if fn not in REAL_EXCEL_FUNCTIONS:
                                FAILURES.append(
                                    f"{path.name}!{ws.title}!{cell.coordinate}: "
                                    f"'{fn}' is not a known Excel function")
    print(f"  {len(seen)} distinct functions used, all recognised"
          if not FAILURES else f"  {len(seen)} distinct functions used")

    # The workbench ships preloaded with worked examples from the tutorials.
    # Verify the arithmetic those sheets will produce.
    section("Workbench preloaded examples")
    check("Workbench one-sample z", (10.9 - 12) / (3.5 / math.sqrt(100)), -3.14, 5e-3)
    check("Workbench two-sample z", (3.9 - 4.8) / math.sqrt(2.1**2 / 50 + 1.9**2 / 40),
          -2.13, 5e-3)
    sp = math.sqrt((63 * 49**2 + 48 * 49**2) / 111)
    check("Workbench two-sample t pooled sd", sp, 49, 1e-9)
    check("Workbench two-sample t", (144 - 121) / (sp * math.sqrt(1 / 64 + 1 / 49)),
          2.473, 5e-3)
    check("Workbench paired t", -8.4 / (4.99 / math.sqrt(30)), -9.22, 0.02)
    check("Workbench chi2 critical", CHISQ_INV_RT(0.05, (2 - 1) * (3 - 1)), 5.99, 5e-3)
    check("Workbench regression t", 0.164 / (math.sqrt(538.6 / 8) / math.sqrt(107702.5)),
          6.56, 0.02)


# ---------------------------------------------------------------------------
# Layer 3 -- YAML content
# ---------------------------------------------------------------------------

def _tolerance(tol, answer, default=5e-3):
    """Mirror the browser's tolerance rules: {abs}, {rel} or a bare number,
    with rel measured against the expected answer."""
    if tol is None:
        return default
    if isinstance(tol, (int, float)):
        return float(tol)
    allowed = float(tol.get("abs", 0) or 0)
    if tol.get("rel") and answer is not None:
        try:
            allowed = max(allowed, abs(float(answer)) * float(tol["rel"]))
        except (TypeError, ValueError):
            pass
    return allowed if allowed else default


def verify_content(ns: dict) -> None:
    section("Question content")
    try:
        import yaml
    except ImportError:
        print("  PyYAML not installed, skipping content checks")
        return

    files = sorted(CONTENT_DIR.glob("tutorial-*.yml"))
    if not files:
        print("  no tutorial YAML files yet")
        return

    seen_ids: dict[str, str] = {}
    n_steps = n_verified = 0

    helpers = dict(
        np=np, math=math, stats=stats,
        stdev_s=stdev_s, var_s=var_s, skew_excel=skew_excel,
        quartile_inc=quartile_inc, mode_sngl=mode_sngl,
        NORM_DIST=NORM_DIST, NORM_INV=NORM_INV, NORM_S_DIST=NORM_S_DIST,
        NORM_S_INV=NORM_S_INV, BINOM_DIST=BINOM_DIST, POISSON_DIST=POISSON_DIST,
        T_INV=T_INV, T_INV_2T=T_INV_2T, T_DIST_2T=T_DIST_2T, T_DIST_RT=T_DIST_RT,
        CHISQ_INV_RT=CHISQ_INV_RT, CHISQ_DIST_RT=CHISQ_DIST_RT,
    )
    # A deliberately small builtins surface: enough to write a readable
    # one-line check, not enough for a content file to reach the filesystem.
    safe_builtins = {
        "abs": abs, "float": float, "int": int, "len": len, "list": list,
        "max": max, "min": min, "pow": pow, "range": range, "round": round,
        "sorted": sorted, "sum": sum, "zip": zip,
    }
    env = {**helpers, **ns}

    for path in files:
        doc = yaml.safe_load(path.read_text())
        for question in doc.get("questions", []):
            qid = question.get("id")
            if not qid:
                FAILURES.append(f"{path.name}: a question is missing an id")
                continue
            if qid in seen_ids:
                FAILURES.append(f"{path.name}: duplicate question id {qid!r} "
                                f"(also in {seen_ids[qid]})")
            seen_ids[qid] = path.name

            for i, step in enumerate(question.get("steps", [])):
                n_steps += 1
                where = f"{path.name} {qid} step {i}"
                kind = step.get("kind")
                if not kind:
                    FAILURES.append(f"{where}: missing kind")
                    continue

                if kind in ("choice", "multi"):
                    opts = step.get("options", [])
                    correct = [o for o in opts if o.get("correct")]
                    if kind == "choice" and len(correct) != 1:
                        FAILURES.append(f"{where}: choice needs exactly one "
                                        f"correct option, found {len(correct)}")
                    if kind == "multi" and not correct:
                        FAILURES.append(f"{where}: multi needs a correct option")
                    for o in opts:
                        if not o.get("correct") and not o.get("feedback"):
                            FAILURES.append(f"{where}: distractor "
                                            f"{o.get('value')!r} has no feedback")

                if kind in ("numeric", "interval"):
                    if "tol" not in step:
                        FAILURES.append(f"{where}: numeric step has no tolerance")
                    for trap in step.get("traps", []):
                        if not trap.get("feedback"):
                            FAILURES.append(f"{where}: trap {trap.get('value')!r} "
                                            f"has no feedback")

                expr = step.get("verify")
                if expr is not None:
                    n_verified += 1
                    try:
                        # YAML happily turns a bare `verify: 7` into an int
                        got = eval(str(expr), {"__builtins__": safe_builtins}, env)
                    except Exception as exc:  # noqa: BLE001
                        FAILURES.append(f"{where}: verify expression failed: {exc}")
                        continue
                    atol = _tolerance(step.get("tol"), step.get("answer"))
                    if kind == "interval":
                        want = [step.get("lower"), step.get("upper")]
                        for label, g, wv in zip(("lower", "upper"), got, want):
                            check(f"{where} {label}", g, wv, atol)
                    elif kind == "table":
                        cells = step.get("cells", [])
                        got = list(got)
                        if len(got) != len(cells):
                            FAILURES.append(
                                f"{where}: verify returned {len(got)} values "
                                f"but the table has {len(cells)} cells")
                            continue
                        for cell, g in zip(cells, got):
                            catol = _tolerance(cell.get("tol", step.get("tol")),
                                               cell.get("answer"), atol)
                            check(f"{where} cell {cell.get('key')}", g,
                                  cell.get("answer"), catol)
                    else:
                        check(f"{where}", got, step.get("answer"), atol)

    print(f"  {len(seen_ids)} questions, {n_steps} steps, "
          f"{n_verified} independently recomputed")


# ---------------------------------------------------------------------------

def main() -> int:
    travel = np.array(load("travel-times.csv")["travel_time_min"])
    junction = load("junction-waiting.csv")
    reaction = load("reaction-distance.csv")

    ns = dict(
        travel=travel,
        travel_no_outlier=np.sort(travel)[:-1],
        fuel=np.array(load("fuel-consumption.csv")["fuel_l_per_100km"]),
        hgv=np.array(load("hgv-weights.csv")["weight_tonnes"]),
        junction_before=np.array(junction["before_s"]),
        junction_after=np.array(junction["after_s"]),
        junction_diff=np.array(junction["after_s"]) - np.array(junction["before_s"]),
        warning=np.array(reaction["warning_distance_m"]),
        reaction=np.array(reaction["reaction_distance_m"]),
    )

    verify_datasets(ns)
    verify_tutorial_2()
    verify_tutorial_3(ns)
    verify_tutorial_4(ns)
    verify_tutorial_5(ns)
    lint_excel_formulas()
    verify_content(ns)

    print()
    if FAILURES:
        print(f"FAILED -- {len(FAILURES)} of {CHECKS} checks did not pass:\n")
        for f in FAILURES:
            print(f"  x {f}")
        return 1
    print(f"All {CHECKS} checks passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
