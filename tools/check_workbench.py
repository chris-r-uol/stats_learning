"""
Evaluate the Hypothesis Test Workbench without Excel.

Linting function *names* catches `CHI.INV.RT`, but it cannot catch a formula
that points at the wrong cell -- and that failure mode is worse, because the
sheet returns a plausible number instead of an error. So this module implements
just enough of Excel's formula language to actually evaluate every sheet and
compare the results against the worked examples from the tutorials.

Run:  python tools/check_workbench.py
"""

from __future__ import annotations

import math
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from scipy import stats

ROOT = Path(__file__).resolve().parents[1]
WB = ROOT / "workbooks" / "hypothesis-test-workbench.xlsx"


# Excel functions used by the workbench, mapped to Python.
FUNCS = {
    "SQRT": math.sqrt,
    "ABS": abs,
    "SUM": lambda *a: sum(a),
    "IF": lambda c, a, b: a if c else b,
    "NORM.S.INV": lambda p: float(stats.norm.ppf(p)),
    "NORM.S.DIST": lambda z, cum: float(stats.norm.cdf(z) if cum else stats.norm.pdf(z)),
    "NORM.INV": lambda p, m, s: float(stats.norm.ppf(p, m, s)),
    "NORM.DIST": lambda x, m, s, cum: float(stats.norm.cdf(x, m, s) if cum
                                            else stats.norm.pdf(x, m, s)),
    "T.INV": lambda p, df: float(stats.t.ppf(p, df)),
    "T.INV.2T": lambda a, df: float(stats.t.ppf(1 - a / 2, df)),
    "T.DIST.2T": lambda t, df: float(2 * stats.t.sf(abs(t), df)),
    "T.DIST.RT": lambda t, df: float(stats.t.sf(t, df)),
    "CHISQ.INV.RT": lambda a, df: float(stats.chi2.isf(a, df)),
    "CHISQ.DIST.RT": lambda x, df: float(stats.chi2.sf(x, df)),
}

CELL_RE = re.compile(r"\b([A-Z]{1,2}\d{1,4})\b")


def translate(formula: str) -> str:
    """Turn an Excel formula into an equivalent Python expression."""
    src = formula.lstrip("=")
    placeholders: dict[str, str] = {}

    def stash(text: str) -> str:
        key = f"__P{len(placeholders)}__"
        placeholders[key] = text
        return key

    # String literals go first. Their contents must survive untouched -- a
    # verdict like "Reject H0" contains something that looks exactly like a
    # cell reference, and substituting into it would corrupt the output.
    src = re.sub(r'"[^"]*"', lambda m: stash(m.group(0)), src)

    # Function names next: they contain dots, and some end in digits
    # (T.INV.2T), which would otherwise confuse the cell-reference pattern.
    for name in sorted(FUNCS, key=len, reverse=True):
        src = src.replace(name + "(", stash("F['" + name + "']") + "(")

    src = re.sub(r"\bTRUE\b", "True", src)
    src = re.sub(r"\bFALSE\b", "False", src)

    src = CELL_RE.sub(lambda m: f"C('{m.group(1)}')", src)

    src = src.replace("^", "**").replace("<>", "!=")
    # Single '=' is comparison in Excel; '==', '<=' and '>=' must survive.
    src = re.sub(r"(?<![<>=!])=(?!=)", "==", src)

    for key, val in placeholders.items():
        src = src.replace(key, val)
    return src


def evaluate_sheet(ws) -> dict[str, object]:
    raw = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                raw[cell.coordinate] = cell.value

    cache: dict[str, object] = {}
    resolving: set[str] = set()

    def C(ref: str):
        if ref in cache:
            return cache[ref]
        if ref in resolving:
            raise ValueError(f"circular reference at {ref}")
        val = raw.get(ref)
        if isinstance(val, str) and val.startswith("="):
            resolving.add(ref)
            try:
                out = eval(translate(val), {"__builtins__": {}},
                           {"C": C, "F": FUNCS})
            finally:
                resolving.discard(ref)
            cache[ref] = out
            return out
        cache[ref] = val
        return val

    for ref in list(raw):
        C(ref)
    return cache


def label_map(ws) -> dict[str, str]:
    """Map the label in column A to the coordinate of the value in column B."""
    out = {}
    for row in ws.iter_rows(min_col=1, max_col=1):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip():
                out[cell.value.strip()] = f"B{cell.row}"
    return out


# (sheet, label, expected, tolerance) -- the worked examples the sheets ship with
EXPECTED = [
    ("One-Sample Z", "Standard error", 0.35, 5e-4),
    ("One-Sample Z", "Test statistic z", -3.142, 5e-3),
    ("One-Sample Z", "Critical value", 1.645, 5e-3),
    ("One-Sample Z", "p-value", 0.000838, 5e-5),
    ("One-Sample Z", "By critical value", "Reject H0", None),
    ("One-Sample Z", "By p-value", "Reject H0", None),
    ("One-Sample Z", "Routes agree?", "yes", None),

    ("One-Sample T", "Degrees of freedom", 29, 0),
    ("One-Sample T", "Test statistic t", 2.739, 5e-3),
    ("One-Sample T", "Critical value", 2.045, 5e-3),
    ("One-Sample T", "Routes agree?", "yes", None),

    ("Two-Sample Z", "Standard error of difference", 0.4224, 5e-4),
    ("Two-Sample Z", "Test statistic z", -2.131, 5e-3),
    ("Two-Sample Z", "Critical value", 1.96, 5e-3),
    ("Two-Sample Z", "p-value", 0.0331, 5e-4),
    ("Two-Sample Z", "Routes agree?", "yes", None),

    ("Two-Sample T", "Pooled sd", 49.0, 1e-6),
    ("Two-Sample T", "Degrees of freedom", 111, 0),
    ("Two-Sample T", "Test statistic t", 2.4727, 5e-3),
    ("Two-Sample T", "Critical value", 1.9816, 5e-3),
    ("Two-Sample T", "Routes agree?", "yes", None),

    ("Paired T", "Degrees of freedom", 29, 0),
    ("Paired T", "Test statistic t", -9.2185, 5e-3),
    ("Paired T", "Critical value", 2.0452, 5e-3),
    ("Paired T", "Routes agree?", "yes", None),

    ("Chi-Squared", "Degrees of freedom", 2, 0),
    ("Chi-Squared", "Critical value", 5.9915, 5e-3),
    ("Chi-Squared", "p-value", 0.1287, 5e-3),
    ("Chi-Squared", "By critical value", "Do not reject H0", None),
    ("Chi-Squared", "Routes agree?", "yes", None),

    ("Regression Coeff T", "Residual standard error", 8.2043, 5e-3),
    ("Regression Coeff T", "Standard error of coefficient", 0.02500, 5e-4),
    ("Regression Coeff T", "Degrees of freedom", 8, 0),
    ("Regression Coeff T", "Test statistic t", 6.5606, 5e-2),
    ("Regression Coeff T", "Critical value", 2.306, 5e-3),
    ("Regression Coeff T", "Routes agree?", "yes", None),
]


def main() -> int:
    wb = load_workbook(WB)
    failures = []
    sheets = {}

    for ws in wb.worksheets:
        if ws.title == "Start Here":
            continue
        try:
            sheets[ws.title] = (evaluate_sheet(ws), label_map(ws))
        except Exception as exc:  # noqa: BLE001
            failures.append(f"{ws.title}: could not evaluate -- {exc}")

    print("Evaluating the workbench without Excel")
    print("-" * 38)

    for sheet, label, want, tol in EXPECTED:
        if sheet not in sheets:
            failures.append(f"{sheet}: sheet missing")
            continue
        cache, labels = sheets[sheet]
        ref = labels.get(label)
        if ref is None:
            failures.append(f"{sheet}: no row labelled {label!r}")
            continue
        got = cache.get(ref)
        if tol is None:
            ok = got == want
        else:
            try:
                ok = abs(float(got) - float(want)) <= tol
            except (TypeError, ValueError):
                ok = False
        if not ok:
            failures.append(f"{sheet} / {label} [{ref}]: got {got!r}, expected {want!r}")

    checked = len(EXPECTED)
    print(f"  {len(sheets)} sheets evaluated, {checked} values checked")

    if failures:
        print(f"\nFAILED -- {len(failures)} problems:\n")
        for f in failures:
            print(f"  x {f}")
        return 1
    print("\nAll workbench formulas evaluate correctly.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
